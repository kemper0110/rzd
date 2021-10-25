from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, Color, NamedStyle
from openpyxl import Workbook


class excel_writer(object):
    #bottom_border = Border(bottom = Side(style = 'medium'))
    #bottom_right_border = Border(bottom = Side(style = 'medium'), right  = Side(style = 'medium'))
    #left_right_border = Border(right  = Side(style = 'medium'), left = Side(style = 'medium'))
    full_border = Border(bottom = Side(style = 'medium'), left = Side(style = 'medium'), right  = Side(style = 'medium'), top = Side(style = 'medium'))
    grey_fill = PatternFill(start_color = 'C8C8C7', end_color='C8C8C7', fill_type='solid')
    red_fill = PatternFill(start_color= 'E21A1A', end_color = 'E21A1A', fill_type= 'solid')
    white_font = Font(color = 'FFFFFF')

    start_posx = 3
    start_posy = 5

    # слишком большая функция
    def write(self, routes, param_keys, old, new):
        self.wb = Workbook()

        header_style = NamedStyle(name="header")
        header_style.border = self.full_border
        header_style.fill = self.red_fill
        header_style.font = self.white_font
        self.wb.add_named_style(header_style)

        regular_style = NamedStyle(name = "regular")
        regular_style.fill = self.grey_fill
        self.wb.add_named_style(regular_style)

        ws = self.wb.active
        posx, posy = self.start_posx, self.start_posy
        
        train_count = 0
        inaccuracy_count = 0
        error_count = 0

        for route in routes:
            route_key = route[0] + ' ' + route[1] + ' ' + route[2]
            for old_train, new_train in zip(old[route_key]['trains_order'], new[route_key]['trains_order']):
                # Сочи -> Москва 21.21.2021 на новом сайте отсутствует 1 маршрут
                # вставка поезда с None значениями, если он отсутствует
                # выглядит ужасно, но вроде работает
                if old_train != new_train:
                    old_trains = old[route_key]['trains_order']
                    new_trains = new[route_key]['trains_order']
                    if len(old_trains) > len(new_trains):
                        for i, train in enumerate(old_trains):
                            if train not in new_trains:
                                new_trains.insert(i, train)
                                new[route_key]['route_data'][train] = dict().fromkeys(param_keys, None)
                    if len(new_trains) > len(old_trains):
                        for i, train in enumerate(new_trains):
                            if train not in old_trains:
                                old_trains.insert(i, train)
                                old[route_key]['route_data'][train] = dict().fromkeys(param_keys, None)

                train = old_train
                train_count += 1
                for i, value in enumerate([route[0] + '->' + route[1] + ' ' + route[2], 'pass.rzd.ru', 'rzd.ru', 'совпадение']) : 
                    ws.cell(row = posy, column = posx - 1 + i, value = value).style = 'header'
                for i in range(2):
                    ws.cell(row = posy + 1, column = posx + i, value = train).style = 'regular'
                # вывод train_id, которые всегда совпадают
                ws.cell(row = posy + 1, column = posx - 1, value = 'train_id').style = 'header'
                ws.cell(row = posy + 1, column = posx + 2, value = '+').style = 'regular'
                for i, param_key in enumerate(param_keys):
                    ws.cell(row = posy + 2 + i, column = posx - 1, value = param_key).style = 'header'
                    ws.cell(row = posy + 2 + i, column = posx, value = old[route_key]["route_data"][train][param_key]).style = 'regular'
                    ws.cell(row = posy + 2 + i, column = posx + 1, value = new[route_key]["route_data"][train][param_key]).style = 'regular'
                    state = '+'
                    if old[route_key]["route_data"][train][param_key] != new[route_key]["route_data"][train][param_key]:
                        # жестко привязано к этим названиям, иначе малейшее отличие будет отмечаться ошибкой
                        if param_key == "station_to" or param_key == "station_from":    
                            state = '*'
                            inaccuracy_count += 1
                        else:
                            state = '-'
                            error_count += 1
                    ws.cell(row = posy + 2 + i, column = posx + 2, value = state).style = 'regular'
                posx += 5
            posx = self.start_posx
            posy += len(param_keys) + 3
        
        ws.cell(column = 1, row = 1, value = f"Количество поездов - {train_count}").style = 'header'
        ws.cell(column = 1, row = 1 + 1, value = f"Количество неточностей - {inaccuracy_count}").style = 'header'
        ws.cell(column = 1, row = 1 + 2, value = f"Количество ошибок - {error_count}").style = 'header'
        

        # width auto_size
        def as_text(value):
            if value is None:
                return ""
            return str(value)
        for column_cells in ws.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        return self

    def save(self):
        from datetime import datetime
        dname = 'out ' + datetime.now().strftime("%m %d %Y, %H %M %S") + '.xlsx'
        self.wb.save(dname)